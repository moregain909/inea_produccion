#from dataclasses import dataclass, field
import copy
import os, sys
from typing import List, Dict, Union
#from decouple import AutoConfig, config, UndefinedValueError
from dotenv import *
#import httpx
from openpyxl import load_workbook, Workbook

#from gbp import proveedor_del_sku, gbp_get_sku_publis
#from recolectar_tipos_envio_ml import (ml_aut, Tienda, get_items_ids, Costos_Envio, Precios, \
#                                       Round_Costos_Envio, Proveedor, Publicacion, SHIPMENT_TYPES_CATALOG, \
#                                        Coeficiente_Envio, get_publis_gbp, get_proveedor_sku)

path2root = os.path.join(os.path.dirname(__file__), "..")
sys.path.append(path2root)

from costos_envio import LISTA_TIENDAS, ALIAS_PROVEEDORES_CON_COSTO_ENVIO, MS_SHIPMENT_MULTIPLIER, \
    MAP_PROVEEDORES, PUBLIS_GBP_SOURCE_FILE, ARTICULOS_GBP_EXTENDIDA, \
    GBP_UPDATER_FILE, GBP_UPDATER_FILE_TITLE_CELLS, GBP_UPDATER_FILE_SHEET_TITLE, GBP_UPDATER_FILE_COLUMNS_WIDTH, \
    DATA_DIR_REL_PATH, DATA_DIR_NAME, SHIPMENT_TYPES_CATALOG

from costos_envio import get_items_ids, Coeficiente_Envio, \
    Proveedor, Shipment_Option, Costos_Envio, Round_Costos_Envio, Precios, Publicacion, \
    get_proveedor_sku, get_publis_gbp, costos_envio_colecta_publi_json, costos_envio_colecta_publi, \
    tiene_mercado_envio, save_to_excel, path2data



from auth import Tienda, format_tienda

### ACTUALIZADOR COSTO DE ENVIO PUBLICACIONES EN GBP


if __name__ == '__main__':
    
    #   CREA LISTA CON PROVEEDORES CON COSTO DE ENVIO EN GBP

    proveedores = [Proveedor(alias=a, map_proveedores=MAP_PROVEEDORES) for a in ALIAS_PROVEEDORES_CON_COSTO_ENVIO]

    #def lista_coeficientes(umbrales: Dict) -> List[Coeficiente_Envio]:
    #    return [Coeficiente_Envio(umbral=k, coeficiente=v) for k, v in umbrales.items()]

    #   AGREGA COEFICIENTES PARA ENVIO GBP A CADA PROVEEDOR
    for p in proveedores:
        p.get_coeficientes_envio()    

    #ids_proveedores_con_costo_envio = [proveedor.pid for proveedor in proveedores]

    #   GENERA TOKENS Y CLIENTS PARA CADA TIENDA
    #   convierte lista de tiendas en diccionario con credenciales con formato {tienda: Tienda()}
    tiendas_dict = {t: Tienda(name=t) for t in LISTA_TIENDAS}
    tiendas = tiendas_dict



    #   GENERA LISTA PUBLIS_ML CON IDS
    print('Generando lista publis ML con IDs')
    publis_ml = []
    for t in tiendas.keys():
        publis_tienda = get_items_ids(store=t, token=tiendas[t].token, client=tiendas[t].client, status="active", offset=0, limit=50)
        publis_ml.extend([Publicacion(tienda=t, item_id=pt) for pt in publis_tienda])
    #print(f'{len(publis_ml)} en publis_ml')

    #   GENERA LISTA PUBLIS_GBP (SIN PROVEEDOR)
    print('Generando lista publis GBP')
    publis_gbp = get_publis_gbp(sku=True, proveedor=False, tienda=True, costo_envio=True, precio=True, \
                                DATA_DIR_REL_PATH=DATA_DIR_REL_PATH, DATA_DIR_NAME=DATA_DIR_NAME, PUBLIS_GBP_SOURCE_FILE=PUBLIS_GBP_SOURCE_FILE)
    #print(f'{len(publis_gbp)} en publis_gbp')

    #   FILTRA PUBLIS GBP ACTIVAS (QUITA LAS QUE NO ESTAN EN PUBLIS_ML)
    #   AGREGA SKU a PUBLIS ML
    print(f'Filtrando publis GBP activas y agregando SKUs')
    for pgbp in publis_gbp[:]:
        found = False
        pgbp.tienda = format_tienda(pgbp.tienda)
        for pml in publis_ml:
            if pgbp.item_id == pml.item_id:
                found = True
                pml.sku = pgbp.sku
                break
        if not found:
            publis_gbp.remove(pgbp)
    #print(f'{len(publis_gbp)} en publis_gbp')    


    #   AGREGA PROVEEDOR A LAS PUBLIS GBP

    print(f'Agregando proveedor a publis GBP')
    #os.chdir(os.path.dirname(os.path.abspath(__file__)))
    # os.chdir("C:/!PYTHON/ML")

    filename = path2data(DATA_DIR_REL_PATH=DATA_DIR_REL_PATH, DATA_DIR_NAME=DATA_DIR_NAME, data_filename=ARTICULOS_GBP_EXTENDIDA)
    articulos_workbook = load_workbook(filename=filename)
    articulos_sheet = articulos_workbook.active

    #print(f'ARTICULOS_GBP_EXTENDIDA: {filename}')
    #print(f'Columnas de Artículos GBP Extendida: {articulos_sheet.max_row}')
    for pgbp in publis_gbp:
        pgbp.proveedor = get_proveedor_sku(pgbp.sku, MAP_PROVEEDORES, Articulos_GBP_extendida_sheet=articulos_sheet)
        pgbp.proveedor.map_proveedores = MAP_PROVEEDORES
        pgbp.proveedor.get_coeficientes_envio()
    #print(f'{len(publis_gbp)} en publis_gbp')

    #   COMPLETA DATOS DE PUBLIS ML
    #   FILTRA POR PROVEEDORES CON ENVIO 
    #   DESCARTA LOS LINKS DE COBRO DE MP QUE ML LISTA COMO PUBLICACIONES
    print(f'Completando datos de publis ML')
    not_found = []
    nueva_publis_ml = []
    for pml in publis_ml:
        found = False

        for pgbp in publis_gbp:
            if pml.item_id == pgbp.item_id:
                
                #print(f'Acá detecta la publi ml {pml.item_id} en publis_gbp')
                if pgbp.proveedor.alias in ALIAS_PROVEEDORES_CON_COSTO_ENVIO:
                    nueva_publi = pml
                    nueva_publi.sku = pgbp.sku
                    nueva_publi.proveedor = pgbp.proveedor
                    nueva_publi.precios = pgbp.precios
                    nueva_publi.costos_envio = Costos_Envio()
                    #print(f'Encuentra el proveedor {pgbp.proveedor.alias} y completa los datos. La nueva_publi ueda así:\n{nueva_publi}\n')
                    nueva_publis_ml.append(copy.deepcopy(nueva_publi))
                    #print(f'nueva_publis_ml queda con {len(nueva_publis_ml)} publicaciones')
                #else:
                #    print(f'Pero no encuentra el proveedor {pgbp.proveedor.alias} en ALIAS_PROVEEDORES_CON_COSTO_ENVIO')
                #print()
                found= True
                break
        if not found:
            not_found.append(pml)

    publis_ml = nueva_publis_ml
    #print(f'{len(publis_ml)} en publis_ml')

    #   PENDIENTE: ACÁ DEBERÍA CHEKEAR SI LAS NOT_FOUND (  "channels": ["mp-merchants", "mp-link"] OR "domain_id": "MLA-MERCADO_POINT_SUPPLIES")
    #   PARA CONFIRMAR QUE NO DESCARTE PUBLICACIONES EN VEZ DE LINKS DE PAGO


    #   TRAE COSTOS DE ENVIO ML 
    #   esto podría hacerse asincrónico para bajar el tiempo de ejecución (>3 min)

    print(f'Trayendo costos de envio ML')
    for i, p in enumerate(publis_ml[:]):
        #   SHIPMENT_TYPES_CATALOG
        item_json = p.get_item_shipment_options_json(token=tiendas[p.tienda.lower()].token, client=tiendas[p.tienda.lower()].client)
        
        if item_json:
            #print(i, p.item_id)
            p.shipment_options = p.get_item_free_shipment_options(item_json, include_prices=True, shipment_types_catalog=SHIPMENT_TYPES_CATALOG)
            #print(f'shipment options: {p.shipment_options}')
            p.set_reference_shipment_costs(ms_shipment_multiplier=MS_SHIPMENT_MULTIPLIER)
            


    #   CALCULA COSTOS ENVIO GBP EN PUBLIS ML
    print(f'Calculando costos envio GBP en publis ML')
    for i, p in enumerate(publis_ml):
        p.set_coeficiente_envio_gbp()
        p.set_costo_envio_gbp()
        #print(f'Envío GBP POST: {i} {p.item_id} {p.precios.venta_gbp} {p.coeficiente_envio_gbp} {p.costos_envio.gbp}')
    #print(f'{len(publis_ml)} en publis_ml')


    ###   COMPARA COSTO DE ENVIO GBP ENTRE PUBLIS ML Y PUBLIS GBP
    ###   CREA LISTA CON PUBLIS A ACTUALIZAR
    print(f'Comparando costos envio GBP entre publis ML y publis GBP')
    actualizacion_publis = []
    for publiml in publis_ml:
        for publigbp in publis_gbp:
            if publiml.item_id == publigbp.item_id:
                if publiml.costos_envio.gbp != publigbp.costos_envio.gbp and publiml.costos_envio.gbp != None:
                    actualizacion_publis.append(publiml)
                    #print(f'{publiml.item_id} venta {publiml.precios.venta_gbp} - envio gbp orig {publigbp.costos_envio.gbp} - envío gbp nuevo {publiml.costos_envio.gbp}')
    #print(f'Actualizaciones: {len(actualizacion_publis)}\n{actualizacion_publis}')

    ###     GUARDAR LISTA DE PUBLIS A ACTUALIZAR EN EXCEL
    print(f'Escribiendo archivo con actualizaciones')
    #print('\n\n')
    excel_pathname = path2data(GBP_UPDATER_FILE, DATA_DIR_REL_PATH=DATA_DIR_REL_PATH, DATA_DIR_NAME=DATA_DIR_NAME)
    
    #print(f'publis_para_actualizar {actualizacion_publis}\nexcel_pathname {excel_pathname}\n\
    #      gbp_updater_file_title_cells {GBP_UPDATER_FILE_TITLE_CELLS}\ngbp_updater_file_sheet_title {GBP_UPDATER_FILE_SHEET_TITLE}\
    #        gbp_updater_file_columns_width {GBP_UPDATER_FILE_COLUMNS_WIDTH}')
    
    if save_to_excel(publis_para_actualizar = actualizacion_publis, excel_pathname = excel_pathname, \
                  gbp_updater_file_title_cells = GBP_UPDATER_FILE_TITLE_CELLS, gbp_updater_file_sheet_title = GBP_UPDATER_FILE_SHEET_TITLE, \
                    gbp_updater_file_columns_width = GBP_UPDATER_FILE_COLUMNS_WIDTH):
    
        print(f'Actualización de costos de envío finalizada.\nSe generó el archivo {excel_pathname}')
    else:
        print(f'No se pudo generar el archivo {excel_pathname}')