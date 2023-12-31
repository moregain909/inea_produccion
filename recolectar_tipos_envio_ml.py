from dataclasses import dataclass, field, asdict
import os
import copy
from typing import List, Dict, Union
from decouple import AutoConfig, config, UndefinedValueError
from dotenv import *
import httpx
from openpyxl import load_workbook, Workbook
#from openpyxl import load_workbook, Workbook
#from gbp import proveedor_del_sku, gbp_get_sku_publis


# SCRIPT PARA IDENTIFICAR TIPOS DE ENVIO DE ML
# Los tipos de envío catalogados hasta el momento se configuran en SHIPMENT_TYPES_CATALOG. 
# El script reporta los tipos de envío que no estén catalogados.
 
#   CONFIGURA EN QUÉ TIENDAS BUSCA TIPOS DE ENVÍO
LISTA_TIENDAS = ["tecnorium", "celestron", "lenovo"]

# CONFIGURA CATALOGO DE TIPOS DE ENVIO DE ML CON SUS PRIORIDADES PARA CALCULAR EL COSTO DE ENVIO DE REFERENCIA
# Con priority >= 50 se toma directamente el costo de envío, con menos, se promedia el costo de todas las prioridades => 0
SHIPMENT_TYPES_CATALOG =    {510645: {'name': 'Estándar a domicilio',                'shipping_method_type': 'three_days',   'priority': -10}, \
                             510945: {'name': 'Estándar a sucursal de correo',       'shipping_method_type': 'three_days',   'priority': 0}, \
                             511546: {'name': 'Estándar a sucursal de correo',       'shipping_method_type': 'four_days',    'priority': 0}, \
                             504345: {'name': 'Estándar a sucursal de correo',       'shipping_method_type': 'standard',     'priority': 0}, \
                              73330: {'name': 'Express a domicilio',                 'shipping_method_type': 'express',      'priority': 100}, \
                             510545: {'name': 'Express a domicilio',                 'shipping_method_type': 'two_days',     'priority': 90}, \
                             510845: {'name': 'Express a sucursal de correo',        'shipping_method_type': 'two_days',     'priority': 0}, \
                             502845: {'name': 'Express a sucursal de correo',        'shipping_method_type': 'express',      'priority': 0}, \
                             510445: {'name': 'Prioritario a domicilio',             'shipping_method_type': 'next_day',     'priority': -10}, \
                             510745: {'name': 'Prioritario a sucursal de correo',    'shipping_method_type': 'next_day',     'priority': 0}, \
                             514245: {'name': 'Prioritario',                         'shipping_method_type': 'five_days',    'priority': -10}, \
                              73328: {'name': 'Estándar a domicilio',                'shipping_method_type': 'standard',     'priority': 60}, \
                             514345: {'name': 'Prioritario',                         'shipping_method_type': 'six_days',     'priority': 60}, \
                             514245: {'name': 'Prioritario',                         'shipping_method_type': 'five_days',    'priority': -10}, \
                              73328: {'name': 'Estándar a domicilio',                'shipping_method_type': 'standard',     'priority': 60}, \
                             514345: {'name': 'Prioritario',                         'shipping_method_type': 'six_days',     'priority': 60}, \
                             511545: {'name': 'Estándar a domicilio',                'shipping_method_type': 'four_days',    'priority': -10} }


#   AUTENTICACIÓN EN ML

#config = AutoConfig(' ')    # previene un bug the decouple que a veces no encuentra .env desde Jupyter

@dataclass
class Credentials:
    store: str = field(repr=True, default="tecnorium")
    app_id: str = field(repr=True, default=None) # también documentado o referenciado como CLIENT_ID
    client_secret = str, field(repr=True, default=None)
    refresh_token = str, field(repr=True, default=None)
    user_id: str = field(repr=True, default=None)

    def __post_init__(self) -> None:
        self.get_credentials()

    def get_credentials(self) -> bool:
        """Trae credenciales para ML de las variables de entorno.
        Argumento:
            store (str): Alias de la tienda. Tiene que estar mapeado en credentials_map.
        Devolución:
            Dict con credenciales (Ej. )
        """

        # Diccionario que mapea store a keys de credenciales de entorno
        credentials_map = {
            ("tecnorium", "tecnorrum"): ("TEC_CLIENT_ID", "TEC_CLIENT_SECRET", "TEC_REFRESH_TOKEN", "TEC_USER_ID"), \
            ("celestron",): ("CEL_CLIENT_ID", "CEL_CLIENT_SECRET", "CEL_REFRESH_TOKEN", "CEL_USER_ID"), \
            ("lenovo", ): ("LEN_CLIENT_ID", "LEN_CLIENT_SECRET", "LEN_REFRESH_TOKEN", "LEN_USER_ID"), \
            ("iojan", "iojann"): ("IOJAN_CLIENT_ID", "IOJAN_CLIENT_SECRET", "IOJAN_REFRESH_TOKEN", "IOJAN_TEC_USER_ID"), \
            ("test01" ,): ("IOJAN_CLIENT_ID", "IOJAN_CLIENT_SECRET", "TEST01_REFRESH_TOKEN", "TEST01_USER_ID")
        }

        # Trae las credenciales de .env con config
        store = self.store.lower()
        for s, c in credentials_map.items():
            if store in s:
                try:
                    self.app_id = config(c[0])
                    self.client_secret = config(c[1])
                    self.refresh_token = config(c[2])
                    self.user_id = config(c[3])
                    return True
                except UndefinedValueError as e:
                    print(f"{e}")
                    #print("Me está dando UndefinedValueError")
                except Exception as e:
                    print(type(e), e)
                    
                return False
        print(f"No se tienen credenciales de la tienda {store}\n") 
        return False
   
            
def ml_aut(tienda: str = "tecnorium", client: httpx.Client = None)-> Union[str, None]:
    """
    Autentica en ML
    Argumentos:
        tienda (str): Alias de una tienda ML de la que tengamos las credenciales en .env y figure mapeado en credentials_map de get_credentials(). Default=None.
        session (httpx.Client()): Cliente de httpx que puede compartirse con otras funciones. Default=None. Si no se declara, crea uno.
    Devolución:
        Entrega un token (str) para autenticar en ML. Si el server devuelve error, la función devuelve False.
    """
    if not client:
        client = httpx.Client()
    else:
        client = client

    c = Credentials(tienda)
    
    # autentica en ML

    url = "https://api.mercadolibre.com/oauth/token"
    payload=f'grant_type=refresh_token&client_id={c.app_id}&client_secret={c.client_secret}&refresh_token={c.refresh_token}'
    headers = {
      'accept': 'application/json',
      'content-type': 'application/x-www-form-urlencoded'
    }
    response = client.post(url, headers=headers, data=payload)
    j = response.json()
    print(j)

    token = f'Bearer {j["access_token"]}'
    
    if response.status_code == 401:
        print("Token inválido")
        return None
    elif response.status_code == 400:
        print(j["error"], j["error_description"])
        return None
    else:
        return(token)

@dataclass
class Tienda:
    name: str = field(repr=True, default=None)
    token: str = field(repr=None, default=None)
    client: httpx.Client = field(repr=None, default=None)

    def __post_init__(self) -> None:
        self.get_token()
        self.get_client()

    def get_token(self) -> None:
        self.token = ml_aut(self.name)
        
    def get_client(self) -> None:
        self.client = httpx.Client()



def get_items_ids(store="tecnorium", token=None, client=None, status="active", offset=0, limit=50) -> Union[List, None]:
    """Trae listado de publicaciones de ML. Sólo trae item_id de la publicación.
    Argumentos:
        store (str) = Alias de la tienda de ML. Default="tecnorium". 
        token (str) = Token de ML. Default=None. 
        session (httpx.Client()) = Cliente de httpx. Permite compartir una misma sesión con el resto del código. Default=None. 
        status (str) = Filtra publicaciones por estado. Default="active". 
        offset (int) = Punto de partida para traer resultados. Default=0.
        limit (int) = Cantidad de publicaciones a traer por llamada a la api. Se suceden las llamada hasta traer el total de reultados (paging). Default=50 (es el máximo). 
    """

    if not client:
        client = httpx.Client()
    else:
        client = client

    if not token:
        token = ml_aut(store, client=client)

    store_id = Credentials(store=store).user_id
    headers = {
        'Authorization': token
    }
    paging = offset + limit
    items = []

    while offset <= paging:
        url = f"https://api.mercadolibre.com/users/{store_id}/items/search?status={status}&offset={offset}&limit={limit}"

        response = client.get(url, headers=headers)

        j = response.json()

        if response.status_code != 200:
            print("No se pudo listado de publicaciones de {store}")
            print(response.status_code, response.text)
            return None
        else:
            paging = j["paging"]["total"]
            page_items = [item for item in j["results"]]
            items.extend(page_items)
            offset = offset + limit
    return items


@dataclass
class Coeficiente_Envio:
    umbral: int = field(repr=True)
    coeficiente: float = field(repr=True)

@dataclass
class Proveedor:
    MAP_PROVEEDORES =  {"microglobal":  {"name": "MICROGLOBAL ARGENTINA SOCIEDAD",  "pid": "16", "umbrales":   {90000: 1, \
                                                                                                                120000: 0.5}}, \
                        "bowie":        {"name": "Bowie SRL",                       "pid": "79", "umbrales":   {90000: 1, \
                                                                                                                120000: 0.5}}}
    
    alias: str = field(repr=True, default=None)
    pid: str = field(repr=True, default=None)
    name: str = field(repr=True, default=None)
    #map_proveedores: Dict[str, str] = field(repr=False, default=dict)
    coeficientes_envio: List[Coeficiente_Envio] = field(default_factory=list)

    def __post_init__(self) -> None:
        if not self.pid:
            self.get_id()
        if not self.alias:
            self.get_alias()
        elif not self.name:
            self.get_name()

    def get_id(self) -> bool:
        if self.alias:
            for alias, details in self.MAP_PROVEEDORES.items():
                if alias == self.alias:
                    self.pid = details["pid"]
                    return True
            print(f"No encontró ID de proveedor de {self.alias}")
            return False        
        elif self.name:
            for details in self.MAP_PROVEEDORES.values():
                if details["name"] == self.name:
                    self.pid = details["pid"]
                    return True
            print(f"No encontró ID de proveedor de {self.name}")
            return False        
        else:
            print(f'Se necesita alias o name del proveedor para poder traer su ID de GBP')
            return False
        
    def get_name(self):
        if self.alias:
            for alias, details in self.MAP_PROVEEDORES.items():
                if alias == self.alias:
                    self.name = details["name"]
                    return True
            #print(f'No tiene info mapeada del proveedor {self.name}')
            return False

    def get_alias(self):
        if self.name:
            for alias, details in self.MAP_PROVEEDORES.items():
                if details["name"] == self.name:
                    self.alias = alias
                    return True
            #print(f'No tiene info mapeada del proveedor {self}')
            return False
    
    def get_coeficientes_envio(self) -> List[Coeficiente_Envio]:

        if self.alias not in self.MAP_PROVEEDORES.keys():
            #print(f'Proveedor {self.alias} no mapeado en MAP_PROVEEDORES')
            return False
        
        for umbral, coeficiente in self.MAP_PROVEEDORES[self.alias]["umbrales"].items():
            self.coeficientes_envio.append(Coeficiente_Envio(umbral=umbral, coeficiente=coeficiente))
        if self.coeficientes_envio:
            return True
        else:
            print(f"No se encontraron coeficientes de envío para proveedor {self.alias}")
            return False


@dataclass
class Shipment_Option:
    name: str = field(default=None, repr=True)
    display: str = field(default=None, repr=True)
    shipping_method: str = field(default=None, repr=True)
    shipping_method_type: str = field(default=None, repr=True)
    price: int = field(default=None, repr=True)
    shipping_method_id: str = field(default=None, repr=True)
    shipping_method_type_priority: int = field(default=None, repr=True)

    def set_priority(self, shipment_types_catalog: dict) -> bool:
        if shipment_types_catalog is None:
            return False
        else:
            if self.shipping_method_id in shipment_types_catalog.keys():
                self.shipping_method_type_priority = shipment_types_catalog[self.shipping_method_id]["priority"]
                return True
            else:
                self.shipping_method_type_priority = 0
                #print(f'Tipo de envío no catalogado en el SHIPMENT_TYPES_CATALOG.\n{self}')
                return False

@dataclass
class Costos_Envio:
    ml: float = field(repr=True, default=None)
    ms: float = field(repr=True, default=None)
    gbp: float = field(repr=True, default=None)


class Round_Costos_Envio:
    # Redondea costos de envío de ML y MS
    @staticmethod
    def round(costos: Costos_Envio):
        Round_Costos_Envio.round_ml(costos)
        Round_Costos_Envio.round_ms(costos)
        Round_Costos_Envio.round_gbp(costos)

    # Redondea costo de envío ML
    @staticmethod
    def round_ml(costos: Costos_Envio):
        if costos.ml:
            costos.ml = round(costos.ml)

    # Redondea costo de envío MS
    @staticmethod
    def round_ms(costos: Costos_Envio):
        if costos.ms:
            costos.ms = round(costos.ms)

    # Redondea costo de envío GBP
    @staticmethod
    def round_gbp(costos: Costos_Envio):
        if costos.gbp:
            costos.gbp = round(costos.gbp)

@dataclass
class Precios:
    venta_gbp: float = field(repr=True, default=None)

@dataclass
class Publicacion:
    item_id: str = field(repr=True, default=None)
    sku: str = field(repr=True, default=None)
    proveedor: Proveedor = field(repr=True, default=None)
    tienda: str = field(repr=True, default=None)
    costos_envio: Costos_Envio() = field(repr=True, default=None)
    shipment_options: list[Shipment_Option] = field(repr=None, default=None)
    precios: Precios = field(repr=True, default=None)
    coeficiente_envio_gbp: float = field(repr=True, default=None)

    def set_coeficiente_envio_gbp(self) -> float:
        if self.proveedor:
            if not self.proveedor.coeficientes_envio:
                print(f'{self.item_id}: Faltan los coeficientes_envio del proveedor {self.proveedor.alias} para calcular el coeficiente de envío gbp a aplicar')
                return False
            elif not self.precios.venta_gbp:
                print(f'{self.item_id}: Faltan los precios de venta GBP para calcular el coeficiente de envío gbp a aplicar')
                return False
            else:
                umbral_activado = 0
                coeficiente_activado = 0
                for coeficiente in self.proveedor.coeficientes_envio:
                    if coeficiente.umbral > self.precios.venta_gbp and (umbral_activado == 0 or (coeficiente.umbral < umbral_activado)):
                        umbral_activado = copy.deepcopy(coeficiente.umbral)
                        coeficiente_activado = copy.deepcopy(coeficiente.coeficiente)
                
                self.coeficiente_envio_gbp = coeficiente_activado
                #print(f'Coeficiente de envío gbp de {self.coeficiente_envio_gbp} agregado a {self.item_id}')
                if umbral_activado != 0:
                    return True
                else:
                    #print(f'{self.item_id}: No se encontró un umbral activo para calcular el coeficiente de envío gbp a aplicar')
                    return False
        else:
            print(f'{self.item_id}: No se encontró el proveedor para calcular el coeficiente de envío gbp a aplicar')
            return False
            
    def set_costo_envio_gbp(self) -> float:
        if not self.costos_envio.ml:
            print(f'{self.item_id}: Falta el costo de envío ML para calcular el costo de envío GBP')
            return False
        else:
            if not self.coeficiente_envio_gbp:
                print(f'{self.item_id}: No hay coeficientes de envío GBP configurado, se aplica 0')
                coeficiente = 0
            else:
                coeficiente = self.coeficiente_envio_gbp
            self.costos_envio.gbp = self.costos_envio.ml * coeficiente
            Round_Costos_Envio.round(self.costos_envio)
            return True

    def get_item_free_shipment_options(self, item_json, include_prices=False, shipment_types_catalog=None) -> Union[List[Shipment_Option], None]:
        if item_json is None:
            print(f'Falta especificar item_shipment_options_json')
            return None
        else:
            try:
                shipment_options = []
                for o in item_json["options"]:

                    #   chekea que la publicación tenga envío gratis
                    if o["cost"] != 0:
                        print(f'La publicación {self.item_id} no tiene envío gratis')
                        return None

                    option = Shipment_Option(name=o["name"], display=o["display"], shipping_method=o["shipping_method_type"], \
                                              shipping_method_id=o["shipping_method_id"], shipping_method_type=o["shipping_method_type"])
                    if include_prices:
                        option.price = o["list_cost"]

                    option.set_priority(shipment_types_catalog)
                    #if (priority := option.set_priority(shipment_types_catalog)):
                    #    option.shipping_method_type_priority = priority

                    if shipment_options:
                        if option not in shipment_options:
                            #self.shipment_options.append(option)
                            shipment_options.append(option)
                    else:
                        shipment_options = [option]
                return shipment_options

            except KeyError as k:
                print(k)
                print(f'La publicación {self.item_id} no tiene opciones de envío. Json:\n{item_json}')
            except Exception as e:
                print(e)
                print(f'Error trayendo opciones de envío para {self.item_id}. Json:\n{item_json}')
            return None
    
    def get_item_shipment_options_json(self, token=None, client=None, zipcode="1602") -> Union[Dict, None]:
    
        if not client:
            client = httpx.Client()

        if not token:
            if not self.tienda:
                print(f"Se requiere token o tienda para traer json con shipment options de la publicación {self.item_id}")
                return None
            else:
                token = ml_aut(self.tienda)

        url = f'https://api.mercadolibre.com/items/{self.item_id}/shipping_options?zip_code={zipcode}'

        headers = {
            'Authorization': token
        }

        response = client.get(url, headers=headers)
        j = response.json()

        if response.status_code == 404:
            print(f'La publicación {self.item_id} no tiene opciones de envío.')
            return None
        elif response.status_code != 200:
                print(f'No se pudo traer opciones de envío de la publicación {self.item_id}.\nStatus code: {response.status_code} - {response.text}')
                return None    
        else:
             return j

    def get_reference_shipment_costs(self, ms_shipment_multiplier=None) -> Union[Costos_Envio, bool]:
        
        if not ms_shipment_multiplier:
            ms_shipment_multiplier = 2
        
        print(self.costos_envio)
        ml_shipment_cost = self.reference_ml_shipment_cost()
        if not ml_shipment_cost or not self.costos_envio:
            self.costos_envio = Costos_Envio()
            return False                
    
        self.costos_envio.ml = ml_shipment_cost
        self.costos_envio.ms = ml_shipment_cost * ms_shipment_multiplier
        Round_Costos_Envio.round(self.costos_envio)
        return True

    def reference_ml_shipment_cost(self) -> Union[int, None]:
        """ 
        Con priority == 100 se toma directamente ese costo de envío y no se busca más.
        Si no hay priority 100 pero hay varias >= 50, se toma el costo de la más alta.
        Si no hay >= 50, se promedian todos los costos con prioridad >= 0.
        """
        if not self.shipment_options:
            print('No se puede calcular el reference_ml_shipment_cost porque la instancia no tiene shipment_options')
            return None
        
        higher_priority_option = None       # >= 50
        mid_priority_options = []           # >= 0
        low_priority_options = []           # < 0
        
        # selecciona por prioridad
        for o in self.shipment_options:
            if o.shipping_method_type_priority or o.shipping_method_type_priority == 0:
                if o.shipping_method_type_priority >= 100:
                    return o.price                          # devuelve primer costo de opción con prioridad 100
                elif o.shipping_method_type_priority >= 50:
                    if higher_priority_option: 
                        if o.price > higher_priority_option.price:
                            higher_priority_option = o      # selecciona por prioridad mayor a 50
                    else:
                        higher_priority_option = o
                elif o.shipping_method_type_priority > 0:
                    mid_priority_options.append(o)             # selecciona por prioridad mayor a 0
                else:
                    low_priority_options.append(o)          # selecciona por prioridad menor a 0
        if higher_priority_option:
            return higher_priority_option.price         # devuelve costo de opción con mejor prioridad mayor a 50
        elif mid_priority_options:
            return self.shipment_cost_average(shipment_options=mid_priority_options)       # devuelve costo promedio de opciones válidas < 50
        elif low_priority_options:
            return self.shipment_cost_average(shipment_options=low_priority_options)    # devuelve costo promedio de opciones low_priority
        else:
            print("No se encuentran opciones de envío válidas")
            return None

    def shipment_cost_average(self, shipment_options: list[Shipment_Option] = None) -> Union[int, None]:
        """Devuelve el costo de envío promedio de una lista con objetos Shipment_Option, o de las shipment_options de la instancia (default)

        Args:
            shipment_options (list, optional): _description_. Defaults to self.shipment_options.

        Returns:
            Union[int, None]: _description_
        """
        if not shipment_options:
            if not self.shipment_options:
                return None
            else:
                shipment_options = self.shipment_options
        
        # filtra por opciones con prioridad mayor a 0
        average_candidates = [opt.price for opt in shipment_options]
        # saca promedio
        total = sum(p for p in average_candidates)
        average = total / len(average_candidates)
        print(f'El costo promedio de las opciones es: {average}')
        print(shipment_options)
        return average


#   TRAE PROVEEDOR EN UNA PUBLICACION

def get_proveedor_sku(sku, Articulos_GBP_extendida_sheet=None):

    if Articulos_GBP_extendida_sheet:
        sheet = Articulos_GBP_extendida_sheet
    else:
        os.chdir(os.path.dirname(os.path.abspath(__file__)))
        # os.chdir("C:/!PYTHON/ML")
        workbook = load_workbook(filename="Articulos_GBP_extendida.xlsx")
        sheet = workbook.active
    rows = sheet.max_row        

    for x in range(2,rows + 1):
        sku_in_row = sheet["F"+str(x)].value
        if sku_in_row == sku:
            p_id = sheet["T"+str(x)].value.split(" | ", 3)[1]
            pname = sheet["T"+str(x)].value.split(" | ", 3)[0]
            proveedor = Proveedor(pid = p_id, name = pname)
            proveedor.get_coeficientes_envio()
            return proveedor
    return False

#   ARMA LISTA CON PUBLICACIONES GBP

def get_publis_gbp(sku=True, proveedor=False, tienda=False, costo_envio=False, precio=False, Articulos_GBP_ext_sheet=None) -> List[Publicacion]:
    # trae List de publis de GBP
    # estaría bueno hacerla asincrónica ya que demora mucho

    #   Levanta excel Publis_GBP
    # os.chdir("C:\!PYTHON\ML")
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    workbook = load_workbook(filename="data/Publis_GBP.xlsx")
    sheet = workbook.active
    rows = int(sheet.dimensions[5:])

    publis_gbp = []

    for x in range(2,rows + 1):
        publi_gbp = Publicacion(item_id = sheet["D"+str(x)].value)

        if sku or proveedor:
            sku_publi = publi_gbp.sku = sheet["F"+str(x)].value
        if sku: sku_publi
        if tienda: publi_gbp.tienda = sheet["AB"+str(x)].value
        if costo_envio: 
            envio = Costos_Envio(gbp = sheet["R"+str(x)].value)
            publi_gbp.costos_envio = envio
        if precio:
            precio = Precios(venta_gbp = sheet["P"+str(x)].value)
            publi_gbp.precios = precio
        if proveedor:
            publi_gbp.proveedor = get_proveedor_sku(sku_publi, Articulos_GBP_extendida_sheet=Articulos_GBP_ext_sheet)

        publis_gbp.append(copy.deepcopy(publi_gbp))

    return publis_gbp


if __name__ == '__main__':
    
    new_shipment_types = {}

    #   GENERA TOKENS Y CLIENTS PARA CADA TIENDA
    #   convierte lista de tiendas en diccionario con credenciales con formato {tienda: Tienda()}
    tiendas = {t: Tienda(name=t) for t in LISTA_TIENDAS}
    
    #   releva opciones de envío de cada publicación de cada tienda
    for v in tiendas.values():

        #   trae listado de ids de publicaciones de la tienda
        publis_tienda = get_items_ids(store=v.name, token=v.token, client=v.client)

        #   trae opciones de envío de cada publicación de la tienda
        for pt in publis_tienda:
            print(v.name, pt)
            publi = Publicacion(item_id=pt)
            shipment_options_json = publi.get_item_shipment_options_json(token=v.token, client=v.client)
            publi.shipment_options = publi.get_item_free_shipment_options(shipment_options_json, include_prices=True, shipment_types_catalog=SHIPMENT_TYPES_CATALOG)

            precio_referencia = publi.reference_ml_shipment_cost()

            # Si la publicación tiene un tipo de envío no catalogado, compara su costo contra el costo de referencia para asignarle prioridad y lo agrega al dict new_shipment_types
            if publi.shipment_options:
                for so in publi.shipment_options:
                    if so.shipping_method_id not in SHIPMENT_TYPES_CATALOG.keys():
                        if so.shipping_method_id not in new_shipment_types.keys():
                            
                            if so.price == precio_referencia:
                                so.shipping_method_type_priority = 60
                            else:
                                so.shipping_method_type_priority = -10
                            new_shipment_types.update({so.shipping_method_id: {'name': so.name, 'shipping_method_type': so.shipping_method_type, 'priority': so.shipping_method_type_priority}})
                            print(f'Agregado este nuevo shipment type: {so.shipping_method_id} - {so.price} - {so.name} - {so.shipping_method_type} - {so.shipping_method_type_priority} / Precio referencia: {precio_referencia}')
                print()
                
        #   Si new_shipment_types tiene contenido, imprimirlo
    if new_shipment_types != {}:
        print(new_shipment_types)    